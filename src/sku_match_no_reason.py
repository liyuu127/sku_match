import re
import uuid
from io import BytesIO

import pandas as pd
from typing import List, Set, Tuple, Any

import time
import asyncio

from llm_match_no_reason import llm_match_fill
from sku_filter import preprocess_candidate_tokens, process_owner_data_async
from utils import load_excel, save_excel_async, download_image
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from openpyxl.utils import get_column_letter

MAX_CONCURRENCY = 100  # 控制最大并发任务数
LLM_MATCH = True


def pandas_str_to_series(s) -> Any | None:
    """字符串转为Series"""
    # 判断是否已经是 Series
    s = str(s)
    s = s.strip()

    # s为None或""或nan
    if s in ["", "nan", "None", "NaN"]:
        return None

    inner = s[s.find("(") + 1: s.rfind(")")]

    pattern = re.compile(r"(\w+)=('[^']*'|[^,]*)")
    data = {k: v.strip("'") if v.strip() != "nan" else None for k, v in pattern.findall(inner)}

    # 转为 DataFrame
    df = pd.DataFrame([data])
    return df.iloc[0]


def pic_url_fill(llm_match_df: pd.DataFrame):
    """根据top1,top2,top3,llm_match 信息填充图片"""

    # 1. 对df单元格数据进行简单清洗，转为str,填充空值
    llm_match_df['商品ID'] = llm_match_df['商品ID'].fillna('').astype(str).str.strip()
    # llm_match_df['相似商品1'] = llm_match_df['相似商品'].fillna('').astype(str).str.strip()
    # llm_match_df['相似商品1'] = llm_match_df['相似商品1'].fillna('').astype(str).str.strip()
    # llm_match_df['相似商品2'] = llm_match_df['相似商品2'].fillna('').astype(str).str.strip()
    # llm_match_df['相似商品3'] = llm_match_df['相似商品3'].fillna('').astype(str).str.strip()
    # 2. 提取商品名称，获取商品图片信息
    llm_match_df['origin_url'] = llm_match_df['图片'].fillna('').astype(str).str.strip()
    llm_match_df['llm_image_url'] = ''
    llm_match_df['top1_image_url'] = ''
    llm_match_df['top2_image_url'] = ''
    llm_match_df['top3_image_url'] = ''

    for index, row in llm_match_df.iterrows():
        top3 = pandas_str_to_series(row['相似商品3'])
        if top3 is not None:
            llm_match_df.at[index, 'top3_image_url'] = top3['图片']

        top2 = pandas_str_to_series(row['相似商品2'])
        if top2 is not None:
            llm_match_df.at[index, 'top2_image_url'] = top2['图片']

        top1 = pandas_str_to_series(row['相似商品1'])
        if top1 is not None:
            llm_match_df.at[index, 'top1_image_url'] = top1['图片']

        llm_p_row = pandas_str_to_series(row['相似商品'])

        if llm_p_row is not None:
            llm_match_df.at[index, 'llm_image_url'] = llm_p_row['图片']

    llm_match_df['top1_image_url'] = llm_match_df['top1_image_url'].fillna('').astype(str).str.strip()
    llm_match_df['top2_image_url'] = llm_match_df['top2_image_url'].fillna('').astype(str).str.strip()
    llm_match_df['top3_image_url'] = llm_match_df['top3_image_url'].fillna('').astype(str).str.strip()
    llm_match_df['llm_image_url'] = llm_match_df['llm_image_url'].fillna('').astype(str).str.strip()

    return llm_match_df


def pic_download(llm_match_df: pd.DataFrame, output_path):
    wb = Workbook()
    ws = wb.active

    # 写入表头（使用 DataFrame 列名）
    ws.append(llm_match_df.columns.tolist())

    row_idx = 2

    for _, row in llm_match_df.iterrows():
        col_idx = 1

        for col_name in llm_match_df.columns:
            value = str(row[col_name]) if row[col_name] is not None and row[col_name] not in ["", "nan", "None"] else \
                row[col_name]
            col_letter = get_column_letter(col_idx)
            # 如果这一列是 image（你指定的需要下载/展示的图片列）
            if col_name in (
                    'origin_url', 'top1_image_url', 'top2_image_url', 'top3_image_url', 'llm_image_url'):
                img_bytes = download_image(value)

                if img_bytes:
                    try:

                        # 加载图片
                        pil_img = PILImage.open(img_bytes).convert("RGB")  # WebP -> RGB
                        pil_img.thumbnail((120, 120))  # 自动缩放

                        # 保存缩放后的图片
                        output = BytesIO()
                        pil_img.save(output, format="PNG")  # 保存为 PNG
                        output.seek(0)

                        # 插入 Excel
                        img = Image(output)
                        cell_pos = f"{col_letter}{row_idx}"
                        ws.add_image(img, cell_pos)

                        # 行高、列宽自适应
                        ws.row_dimensions[row_idx].height = 90
                        ws.column_dimensions[col_letter].width = 18
                    except Exception as e:
                        print(f"图片解析失败：{e}")
                        ws.cell(row=row_idx, column=col_idx, value=value)
                else:
                    ws.cell(row=row_idx, column=col_idx, value=value)

            else:
                # 普通列写值
                ws.cell(row=row_idx, column=col_idx, value=value)

            col_idx += 1

        row_idx += 1
    wb.save(output_path)
    print(f"数据写入完成！,文件：{output_path}")


async def main():
    print("开始加载数据...")
    start_time = time.time()
    # ele_df = pd.read_csv("../data/elme_sku_small.csv")
    # owner_df = pd.read_csv("../data/meituan_sku_small.csv")

    # owner_df = load_excel("../data/美团-快驿点特价超市(虹桥店)全量商品信息20251109.xlsx").iloc[:1000]
    # target_df = load_excel("../data/美团-邻侣超市（虹桥中心店）全量商品信息20251109.xlsx").iloc[:1000]
    # owner_df = load_excel("../data/美团-快驿点特价超市(虹桥店)全量商品信息20251109.xlsx")
    # target_df = load_excel("../data/美团-邻侣超市（虹桥中心店）全量商品信息20251109.xlsx")
    # owner_df = load_excel("../data/sku_kyd_sampled.xlsx").iloc[:100]
    owner_df = load_excel("../data/top3相似人工标注数据_需要大模型识别.xlsx").iloc[:50]
    # owner_df = load_excel("../output/llm_error_df.xlsx")
    target_df = load_excel("../data/附件2-美团邻侣全量去重商品1109.xlsx")
    owner_df = owner_df.drop_duplicates(subset=['商品ID'])
    target_df = target_df.drop_duplicates(subset=['商品ID'])
    # 打印前几行数据
    print(owner_df.head())
    print(f"待匹配数据加载完成，共{len(owner_df)}条记录")
    print(target_df.head())
    print(f"匹配目标数据加载完成，共{len(target_df)}条记录")

    tokens = await preprocess_candidate_tokens(target_df)
    print("饿了么数据预处理完成")

    print("开始异步处理匹配任务...")
    top1_list, top2_list, top3_list = await process_owner_data_async(owner_df, target_df, tokens)

    owner_df['相似商品1'] = top1_list
    owner_df['相似商品2'] = top2_list
    owner_df['相似商品3'] = top3_list

    suffix = str(uuid.uuid4())
    output_path = "../output/top3相似_qwen3_30B_500_" + suffix + ".xlsx"
    output_path_nopic = "../output/nopic_top3相似_qwen3_30B_500_" + suffix + ".xlsx"
    if LLM_MATCH:
        await llm_match_fill(owner_df, top1_list, top2_list, top3_list)
        pic_url_fill(owner_df)
        await save_excel_async(owner_df, output_path_nopic)
        pic_download(owner_df, output_path)

    # output_path = "../output/top3相似_qwen3_30B_500_" + str(uuid.uuid4()) + ".xlsx"
    # await save_excel_async(owner_df, output_path)

    end_time = time.time()
    print(f"\n处理完成！结果已保存到：{output_path_nopic}")
    print(f"总耗时：{end_time - start_time:.2f}秒")


if __name__ == '__main__':
    asyncio.run(main())
